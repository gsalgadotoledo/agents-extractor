"""Agent chat with file uploads, audio transcription, and agentic data processing."""
from __future__ import annotations

import csv
import io
import json
import operator
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated, TypedDict

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from langchain_anthropic import ChatAnthropic
from langchain_core.messages import BaseMessage, HumanMessage, SystemMessage
from langchain_core.tools import tool
from langgraph.graph import END, StateGraph
from langgraph.prebuilt import ToolNode

from submission_platform.api.dependencies import get_app_settings, get_store
from submission_platform.config import Settings
from submission_platform.domain import submissions
from submission_platform.domain.models import Submission
from submission_platform.infra.json_store import JsonStore
from submission_platform.infra.logging import get_logger

log = get_logger(__name__)
router = APIRouter(prefix="/submissions", tags=["chat"])

_chat_histories: dict[str, list[BaseMessage]] = {}


class _ChatState(TypedDict):
    messages: Annotated[list[BaseMessage], operator.add]


# --- File parsing utilities ---

def _parse_pdf(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text}")
    return "\n\n".join(pages) if pages else "(No text extracted from PDF)"


def _parse_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return "(Empty CSV)"
    # Format as table
    lines = []
    for i, row in enumerate(rows[:100]):  # limit to 100 rows
        lines.append(" | ".join(row))
        if i == 0:
            lines.append("-" * len(lines[0]))
    if len(rows) > 100:
        lines.append(f"... ({len(rows) - 100} more rows)")
    return "\n".join(lines)


def _parse_excel(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 100:
                rows.append(f"... (more rows)")
                break
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
            if i == 0:
                rows.append("-" * max(len(rows[0]), 20))
        parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _parse_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _parse_image_description(data: bytes, filename: str) -> str:
    return f"[Image file: {filename}, {len(data)} bytes. Image content analysis requires vision model — treating as attachment reference.]"


def parse_file(data: bytes, filename: str, content_type: str) -> str:
    """Parse file content into text based on type."""
    lower = filename.lower()
    if lower.endswith(".pdf") or content_type == "application/pdf":
        return _parse_pdf(data)
    elif lower.endswith(".csv") or content_type == "text/csv":
        return _parse_csv(data)
    elif lower.endswith((".xlsx", ".xls")) or "spreadsheet" in content_type:
        return _parse_excel(data)
    elif lower.endswith((".txt", ".md", ".log", ".json", ".xml")):
        return _parse_text(data)
    elif content_type.startswith("text/"):
        return _parse_text(data)
    elif content_type.startswith("image/"):
        return _parse_image_description(data, filename)
    else:
        return f"[File: {filename}, {len(data)} bytes, type: {content_type}]"


# --- Audio transcription ---

async def transcribe_audio(audio_data: bytes, filename: str, settings: Settings) -> str:
    """Transcribe audio using OpenAI Whisper."""
    if not settings.openai_api_key:
        return "(Audio transcription requires OPENAI_API_KEY in .env)"

    import openai
    client = openai.OpenAI(api_key=settings.openai_api_key)

    # Save to temp file (Whisper needs a file)
    suffix = Path(filename).suffix or ".webm"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(audio_data)
        tmp_path = tmp.name

    try:
        with open(tmp_path, "rb") as f:
            transcript = client.audio.transcriptions.create(model="whisper-1", file=f)
        return transcript.text
    finally:
        Path(tmp_path).unlink(missing_ok=True)


# --- Chat agent tools ---

def _make_chat_tools(submission_id: str, store: JsonStore, file_contents: dict[str, str]):
    """Create tools for the chat agent with file processing context."""

    @tool
    def get_extracted_data() -> str:
        """Get the current extracted data for this submission as JSON."""
        sub = store.load("submissions", submission_id, Submission)
        if sub is None or sub.extracted_data is None:
            return "No extracted data available."
        return json.dumps(sub.extracted_data, indent=2)

    @tool
    def update_field(path: str, value: str) -> str:
        """Update a specific field in the extracted data using dot notation.

        Args:
            path: Dot-separated path like "overview.insured_name" or "loss_runs.present" or "facilities.0.city"
            value: The new value. For booleans use "true"/"false", for numbers use digits.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None or sub.extracted_data is None:
            return "Error: no extracted data."

        parts = path.split(".")
        target = sub.extracted_data
        for part in parts[:-1]:
            if part.isdigit():
                idx = int(part)
                if isinstance(target, list) and idx < len(target):
                    target = target[idx]
                else:
                    return f"Error: index {part} out of range in path '{path}'"
            elif isinstance(target, dict):
                if part not in target:
                    target[part] = {}
                target = target[part]
            else:
                return f"Error: invalid path '{path}'"

        last = parts[-1]
        parsed: str | int | float | bool = value
        if value.lower() == "true": parsed = True
        elif value.lower() == "false": parsed = False
        elif value.replace(".", "").replace("-", "").isdigit():
            parsed = int(value) if "." not in value else float(value)

        if isinstance(target, dict):
            target[last] = parsed
        elif isinstance(target, list) and last.isdigit():
            target[int(last)] = parsed
        else:
            return f"Error: cannot set '{path}'"

        store.save("submissions", submission_id, sub)
        return f"Updated {path} = {parsed}"

    @tool
    def replace_section(section: str, data_json: str) -> str:
        """Replace an entire section of the extracted data with new data.

        Args:
            section: Top-level key like "loss_runs", "facilities", "overview", "broker", "coverage", etc.
            data_json: The new data as a JSON string.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None or sub.extracted_data is None:
            return "Error: no extracted data."
        try:
            new_data = json.loads(data_json)
            sub.extracted_data[section] = new_data
            store.save("submissions", submission_id, sub)
            return f"Replaced section '{section}' successfully."
        except json.JSONDecodeError as e:
            return f"Error: invalid JSON — {e}"

    @tool
    def add_to_list(path: str, item_json: str) -> str:
        """Add an item to an array field in the extracted data.

        Args:
            path: Dot path to the array, e.g. "facilities", "claims_history", "loss_runs.periods"
            item_json: The new item as a JSON string.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None or sub.extracted_data is None:
            return "Error: no extracted data."
        try:
            item = json.loads(item_json)
        except json.JSONDecodeError as e:
            return f"Error: invalid JSON — {e}"

        parts = path.split(".")
        target = sub.extracted_data
        for part in parts:
            if part.isdigit():
                target = target[int(part)]
            elif isinstance(target, dict):
                target = target.get(part)
            if target is None:
                return f"Error: path '{path}' not found"

        if not isinstance(target, list):
            return f"Error: '{path}' is not a list"

        target.append(item)
        store.save("submissions", submission_id, sub)
        return f"Added item to {path}. Now has {len(target)} items."

    @tool
    def read_uploaded_file(filename: str) -> str:
        """Read the extracted text content of an uploaded file.

        Args:
            filename: The name of the uploaded file to read.
        """
        if filename in file_contents:
            return f"Content of {filename}:\n\n{file_contents[filename]}"
        available = list(file_contents.keys()) if file_contents else ["(no files uploaded)"]
        return f"File '{filename}' not found. Available: {available}"

    @tool
    def get_email_content() -> str:
        """Get the original email subject and body for context."""
        sub = store.load("submissions", submission_id, Submission)
        if sub is None:
            return "Submission not found."
        return f"SUBJECT: {sub.subject}\nFROM: {sub.broker_email}\n\nBODY:\n{sub.body_text}"

    @tool
    def add_comment(comment: str) -> str:
        """Add a comment/note to the submission timeline visible to all team members.

        Args:
            comment: The comment text to add to the timeline.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None:
            return "Error: submission not found."
        sub.chat_history.append({"role": "system", "content": f"Note: {comment}"})
        sub.updated_at = datetime.now(timezone.utc)
        store.save("submissions", submission_id, sub)
        return f"Comment added to timeline: {comment}"

    @tool
    def change_status(new_status: str) -> str:
        """Change the submission status. Valid statuses: received, ack_sent, extracting, extracted, validated, needs_review, auto_policy_ready, policy_created, completed, failed.

        Args:
            new_status: The target status.
        """
        from submission_platform.domain.models import SubmissionStatus
        sub = store.load("submissions", submission_id, Submission)
        if sub is None:
            return "Error: submission not found."
        try:
            new_s = SubmissionStatus(new_status)
        except ValueError:
            return f"Error: invalid status '{new_status}'"
        old = sub.status.value
        sub.status = new_s
        sub.chat_history.append({"role": "system", "content": f"Status changed: {old} → {new_status}"})
        sub.updated_at = datetime.now(timezone.utc)
        store.save("submissions", submission_id, sub)
        return f"Status changed from {old} to {new_status}"

    @tool
    def send_email_to_broker(subject: str, body: str) -> str:
        """Send an email to the broker and record it in the timeline.

        Args:
            subject: Email subject line.
            body: Email body text.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None:
            return "Error: submission not found."
        # Record the email (actual sending happens async separately)
        sub.sent_emails.append({
            "to": sub.broker_email,
            "subject": subject,
            "body": body,
            "sent_at": datetime.now(timezone.utc).isoformat(),
        })
        sub.chat_history.append({"role": "system", "content": f"Email queued to {sub.broker_email}: {subject}"})
        sub.updated_at = datetime.now(timezone.utc)
        store.save("submissions", submission_id, sub)
        return f"Email queued to {sub.broker_email}. Subject: {subject}"

    @tool
    def assign_to_rep(rep_name: str) -> str:
        """Assign or reassign this submission to a representative.

        Args:
            rep_name: The representative's name (e.g. "Sarah Chen", "Marcus Johnson").
        """
        from submission_platform.domain.assignment import ALL_USERS, REPRESENTATIVES
        rep = next((r for r in REPRESENTATIVES if r["name"].lower() == rep_name.lower()), None)
        if not rep:
            names = [r["name"] for r in REPRESENTATIVES]
            return f"Unknown rep '{rep_name}'. Available: {names}"
        sub = store.load("submissions", submission_id, Submission)
        if sub is None:
            return "Error: submission not found."
        old = ALL_USERS.get(sub.assigned_to, {}).get("name", "nobody")
        sub.assigned_to = rep["id"]
        sub.chat_history.append({"role": "system", "content": f"Reassigned from {old} to {rep['name']}"})
        sub.updated_at = datetime.now(timezone.utc)
        store.save("submissions", submission_id, sub)
        return f"Assigned to {rep['name']}"

    @tool
    def generate_document(doc_type: str) -> str:
        """Generate a PDF document from the submission data. Types: quote, policy, binder, certificate.

        Args:
            doc_type: One of: quote, policy, binder, certificate.
        """
        sub = store.load("submissions", submission_id, Submission)
        if sub is None or sub.extracted_data is None:
            return "Error: no extracted data to generate from."
        sub.chat_history.append({"role": "system", "content": f"Document generation requested: {doc_type}"})
        store.save("submissions", submission_id, sub)
        return f"Document generation for '{doc_type}' has been queued. Use the Documents tab to generate and download."

    @tool
    def draft_email(to: str, subject: str, body: str) -> str:
        """Draft an email for the representative to review before sending. \
Use this instead of send_email_to_broker when the rep asks to compose, write, or draft an email. \
The rep will be able to edit the email before sending it.

        Args:
            to: Recipient email address.
            subject: Email subject line.
            body: The full email body text.
        """
        # Return a special JSON format the frontend will detect and render as an editable card
        draft = json.dumps({
            "__type": "email_draft",
            "to": to,
            "subject": subject,
            "body": body,
            "submission_id": submission_id,
        })
        return f"EMAIL_DRAFT:{draft}"

    @tool
    def get_workflow_rules() -> str:
        """Get the workflow rules that define when status transitions and document generation are allowed."""
        from pathlib import Path
        rules_path = Path(__file__).parent.parent.parent / "domain" / "workflow_rules.md"
        if rules_path.exists():
            return rules_path.read_text()
        return "Workflow rules file not found."

    return [
        get_extracted_data, update_field, replace_section, add_to_list,
        read_uploaded_file, get_email_content,
        add_comment, change_status, send_email_to_broker, draft_email, assign_to_rep,
        generate_document, get_workflow_rules,
    ]


CHAT_SYSTEM_PROMPT = """\
You are the submission processing assistant for an insurance underwriting platform. \
You work alongside representatives to process commercial insurance submissions \
(primarily General Liability) from intake to policy issuance.

CONTEXT:
This platform receives insurance applications via email from brokers. Each submission \
goes through: intake, data extraction, review, quoting, approval, and policy issuance. \
Your role is to help representatives at every step.

THE SUBMISSION LIFECYCLE:
1. Email received from broker with application and attachments
2. AI extracts structured data (insured info, coverage, loss runs, facilities)
3. Representative reviews extracted data, fills gaps, corrects errors
4. Missing information is requested from the broker via email
5. Once complete, submission moves to quoting
6. Quote is generated and sent to broker
7. If accepted, policy is generated after manager approval

TOOLS:
- get_extracted_data / update_field / replace_section / add_to_list: manage submission data
- read_uploaded_file: parse files the rep uploads (PDFs, CSVs, Excel)
- get_email_content: re-read the original broker email
- draft_email: compose an email draft the rep can review and edit before sending (PREFERRED)
- send_email_to_broker: send an email immediately without review (use only when rep says "send it now")
- add_comment: add an internal note to the timeline (visible to all team members)
- change_status: advance the submission through the pipeline
- assign_to_rep: assign or reassign to a different representative
- generate_document: request PDF generation (quote, policy, binder, certificate)
- get_workflow_rules: check what transitions and actions are currently valid

HOW TO BEHAVE:
- Be concise and direct. This is a professional but casual work conversation.
- NEVER use emojis or emoji unicode characters in your responses. No checkmarks (✅), \
  no warning signs (⚠️), no sparkles, no thumbs up, no icons of any kind. \
  Use plain text only. Use markdown formatting (bold, lists, headers) instead of emojis \
  to structure your responses. For example use "Done:" instead of "✅", use "Note:" \
  instead of "⚠️", use "**Next steps:**" instead of "🔜".
- When the rep says "set X to Y" or "the FEIN is 12345", immediately update the field \
  and confirm briefly: "Updated FEIN to 12345."
- When the rep uploads a file, read it thoroughly and extract ALL relevant data. \
  Update the submission with everything you find. Summarize what you extracted.
- When asked what's missing, check get_extracted_data and list specifically what fields \
  are empty or incomplete, and what the rep or broker needs to provide.
- When recording changes, use add_comment to note WHO requested the change. \
  Example: "FEIN updated to 84-1234567 per rep Sarah Chen."
- When asked to write, compose, or draft an email, ALWAYS use draft_email so the rep can review it. \
  Only use send_email_to_broker if the rep explicitly says "send it" or "send it now". \
  Keep the tone business-casual. Be specific about what information is needed and why. \
  Sign emails with the sender name from the current context (see below).
- Before changing status, always check get_workflow_rules to verify the transition is valid.
- Before generating documents, verify the submission has all required fields filled.

WHAT YOU SHOULD NOT DO:
- ABSOLUTELY NO EMOJIS. Not even common ones like checkmarks or arrows. Plain text only.
- When you receive a message marked as [Voice transcription], the text was transcribed from audio \
  and may contain errors. If the transcription seems unclear, contains nonsense words, or the \
  requested action doesn't make sense, ask the user to confirm before making any changes. \
  Example: "I heard 'set the FEIN to eighty four dash one two three four five six seven'. \
  Did you mean 84-1234567? Please confirm and I'll update it."
- Do not make up data. If you don't know a value, say so
- Do not change status without verifying requirements are met
- Do not send emails without the rep explicitly asking

PROACTIVE NEXT STEPS — THIS IS CRITICAL:
After every response, you MUST end with a "Next steps" section that tells the rep \
exactly what should happen next to move this submission forward. Be specific and actionable.

Depending on the current state, suggest things like:
- If data was just extracted: "Next steps: Review the extracted data above. \
  There are 3 missing fields — I can draft an email to the broker requesting them, \
  or you can provide the values here."
- If missing fields were just filled: "Next steps: All required fields are now complete. \
  I can move this to quoting status and generate a quote proposal. Want me to proceed?"
- If a quote was generated: "Next steps: The quote is ready for review. \
  Once you're satisfied, I can send it to the broker and request manager approval."
- If waiting on broker response: "Next steps: We're waiting on the broker to provide \
  loss runs. I can send a follow-up email if it's been more than 2 business days."
- If approved: "Next steps: This submission is approved. I can generate the policy \
  document and send it to the broker to finalize."

Always frame next steps as offers: "I can do X" or "Would you like me to Y?" \
so the rep can simply say yes and you take the action immediately.

Format the next steps clearly using **bold** for the section header.

CURRENT SUBMISSION CONTEXT:
{submission_context}

{file_context}
"""


# --- Chat endpoint ---

@router.post("/{submission_id}/chat")
async def chat_with_agent(
    submission_id: str,
    message: str = Form(""),
    audio: UploadFile | None = File(None),
    files: list[UploadFile] | None = File(None),
    store: JsonStore = Depends(get_store),
    settings: Settings = Depends(get_app_settings),
):
    sub = await submissions.get_submission(submission_id, store=store)
    if sub is None:
        raise HTTPException(status_code=404, detail="Submission not found")

    # Process audio transcription
    user_text = message
    transcription = None
    if audio and audio.filename:
        audio_data = await audio.read()
        if audio_data:
            transcription = await transcribe_audio(audio_data, audio.filename, settings)
            user_text = f"{user_text}\n\n[Voice transcription]: {transcription}" if user_text else f"[Voice transcription]: {transcription}\n\n[NOTE: This was transcribed from audio. If the request seems unclear or doesn't make sense, ask the user to confirm or clarify before making changes.]"

    # Process uploaded files
    file_contents: dict[str, str] = {}
    file_names: list[str] = []
    for f in (files or []):
        if f.filename and f.size and f.size > 0:
            data = await f.read()
            content_type = f.content_type or "application/octet-stream"
            parsed = parse_file(data, f.filename, content_type)
            file_contents[f.filename] = parsed
            file_names.append(f.filename)
            log.info("chat_file_uploaded", filename=f.filename, chars=len(parsed))

            # Also save file to disk for reference
            attach_dir = store._base_dir / "chat_uploads" / submission_id
            attach_dir.mkdir(parents=True, exist_ok=True)
            (attach_dir / f.filename).write_bytes(data)

    # Build file context for the system prompt
    file_context = ""
    if file_names:
        file_context = f"\nThe user has uploaded {len(file_names)} file(s): {', '.join(file_names)}. Use read_uploaded_file to read their contents."
        # Append file info to user message
        user_text += f"\n\n[Uploaded files: {', '.join(file_names)}]"

    if not user_text.strip():
        return {"reply": "Please type a message or upload a file.", "updated_fields": [], "transcription": None, "files_processed": []}

    # Build agent
    tools = _make_chat_tools(submission_id, store, file_contents)

    llm = ChatAnthropic(
        model=settings.extraction_model,
        anthropic_api_key=settings.anthropic_api_key,
        temperature=0,
        max_tokens=4096,
    ).bind_tools(tools)

    tool_node = ToolNode(tools)

    def call_model(state: _ChatState) -> dict:
        return {"messages": [llm.invoke(state["messages"])]}

    def should_continue(state: _ChatState) -> str:
        last = state["messages"][-1]
        if hasattr(last, "tool_calls") and last.tool_calls:
            return "tools"
        return END

    graph = StateGraph(_ChatState)
    graph.add_node("agent", call_model)
    graph.add_node("tools", tool_node)
    graph.set_entry_point("agent")
    graph.add_conditional_edges("agent", should_continue, {"tools": "tools", END: END})
    graph.add_edge("tools", "agent")
    compiled = graph.compile()

    # Load persisted chat history from submission
    sub = store.load("submissions", submission_id, Submission)
    # Build dynamic context about this submission
    from submission_platform.domain.assignment import ALL_USERS
    from submission_platform.domain.personas import get_persona_for_submission
    sub_reload = store.load("submissions", submission_id, Submission)
    assigned_name = ALL_USERS.get(sub_reload.assigned_to, {}).get("name", "Unassigned") if sub_reload else "Unknown"
    persona = get_persona_for_submission(sub_reload, store) if sub_reload else None
    persona_name = persona.email_name if persona else settings.email_from_name
    persona_email = persona.email_address if persona else settings.email_from_address
    persona_tone = persona.tone if persona else "professional"
    persona_personality = persona.personality if persona else "Professional and concise."
    persona_sig = persona.signature.replace("\n", " | ") if persona else settings.email_from_name
    persona_greeting = persona.greeting_style if persona else "Hi,"
    overview = (sub_reload.extracted_data or {}).get("overview", {}) if sub_reload else {}
    missing = (sub_reload.extracted_data or {}).get("missing_fields", []) if sub_reload else []

    submission_context = f"""Submission ID: {submission_id[:8]}
Status: {sub_reload.status.value if sub_reload else 'unknown'}
Insured: {overview.get('insured_name', 'Not yet extracted')}
Broker: {sub_reload.broker_email if sub_reload else 'unknown'}
Assigned to: {assigned_name}
Approved: {'Yes' if sub_reload and sub_reload.approved_by else 'No'}
Missing fields: {len(missing)} ({', '.join(missing[:5])}{'...' if len(missing) > 5 else ''})
Attachments: {len(sub_reload.attachments) if sub_reload else 0}
Confidence in data extraction: {f'{sub_reload.extraction_confidence:.0%}' if sub_reload and sub_reload.extraction_confidence else 'N/A'} (internal metric — do not mention to broker)
Email sender name: {persona_name}
Email sender address: {persona_email}
Persona tone: {persona_tone}
Persona personality: {persona_personality}
Persona signature: {persona_sig}
Persona greeting: {persona_greeting}
When drafting emails, ALWAYS use this persona's tone, personality, greeting style, and signature. Stay in character."""

    system_prompt = CHAT_SYSTEM_PROMPT.format(
        submission_context=submission_context,
        file_context=file_context,
    )
    langchain_history: list[BaseMessage] = [SystemMessage(content=system_prompt)]

    # Rebuild LangChain messages from persisted history (skip system entries, only user/assistant)
    from langchain_core.messages import AIMessage
    for entry in (sub.chat_history if sub else []):
        if entry.get("role") == "user":
            langchain_history.append(HumanMessage(content=entry.get("content", "")))
        elif entry.get("role") == "assistant":
            langchain_history.append(AIMessage(content=entry.get("content", "")))

    langchain_history.append(HumanMessage(content=user_text))

    try:
        result = await compiled.ainvoke({"messages": langchain_history})
    except Exception as e:
        log.error("chat_agent_failed", error=str(e))
        # Persist the error in chat history
        sub = store.load("submissions", submission_id, Submission)
        if sub:
            sub.chat_history.append({"role": "user", "content": user_text, "files": file_names, "transcription": transcription})
            sub.chat_history.append({"role": "assistant", "content": f"Error: {e}"})
            store.save("submissions", submission_id, sub)
        return {
            "reply": f"Agent error: {e}",
            "updated_fields": [],
            "transcription": transcription,
            "files_processed": file_names,
            "tool_steps": [],
        }

    # Extract reply and tool steps
    last_msg = result["messages"][-1]
    reply = last_msg.content if isinstance(last_msg.content, str) else str(last_msg.content)

    updated_fields = []
    tool_steps = []
    email_drafts = []
    for msg in result["messages"]:
        # Track tool calls
        if hasattr(msg, "tool_calls") and msg.tool_calls:
            for tc in msg.tool_calls:
                tool_steps.append({"tool": tc["name"], "args": tc.get("args", {}), "type": "call"})
        # Track tool results
        content = getattr(msg, "content", "")
        if isinstance(content, str):
            if content.startswith("EMAIL_DRAFT:"):
                try:
                    draft_data = json.loads(content.split("EMAIL_DRAFT:", 1)[1])
                    email_drafts.append(draft_data)
                    tool_steps.append({"tool": "draft_email", "result": "Email draft created", "type": "result"})
                except json.JSONDecodeError:
                    pass
            elif content.startswith("Updated "):
                field = content.split("Updated ", 1)[1].split(" = ")[0]
                updated_fields.append(field)
                tool_steps.append({"tool": "update_field", "result": content, "type": "result"})
            elif content.startswith("Replaced section"):
                section = content.split("'")[1] if "'" in content else "unknown"
                updated_fields.append(f"(replaced {section})")
                tool_steps.append({"tool": "replace_section", "result": content, "type": "result"})
            elif content.startswith("Added item"):
                path = content.split("Added item to ", 1)[1].split(".")[0] if "Added item to " in content else "unknown"
                updated_fields.append(f"(added to {path})")
                tool_steps.append({"tool": "add_to_list", "result": content, "type": "result"})

    # Persist chat entry to submission
    sub = store.load("submissions", submission_id, Submission)
    if sub:
        sub.chat_history.append({"role": "user", "content": user_text, "files": file_names, "transcription": transcription})
        sub.chat_history.append({"role": "assistant", "content": reply, "updated_fields": updated_fields, "tool_steps": tool_steps})
        store.save("submissions", submission_id, sub)

    return {
        "reply": reply,
        "updated_fields": updated_fields,
        "transcription": transcription,
        "files_processed": file_names,
        "tool_steps": tool_steps,
        "email_drafts": email_drafts,
    }


@router.get("/{submission_id}/chat")
async def get_chat_history(
    submission_id: str,
    store: JsonStore = Depends(get_store),
):
    """Get persisted chat history for a submission."""
    sub = await submissions.get_submission(submission_id, store=store)
    if sub is None:
        raise HTTPException(status_code=404, detail="Submission not found")
    return {"history": sub.chat_history}


@router.delete("/{submission_id}/chat")
async def clear_chat_history(
    submission_id: str,
    store: JsonStore = Depends(get_store),
):
    sub = await submissions.get_submission(submission_id, store=store)
    if sub:
        sub.chat_history = []
        store.save("submissions", sub.id, sub)
    return {"ok": True}
